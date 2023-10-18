import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def calcular_valor_custo(valor, quantidade, frete, imposto1, imposto2, imposto3):
    return valor * quantidade + frete + (valor * (imposto1/100)) + (valor * (imposto2/100)) + (valor * (imposto3/100))

def cadastrar_produto():
    nome_produto = input("Digite o nome do produto: ")
    valor = float(input("Digite o valor do produto: "))
    quantidade = int(input("Digite a quantidade: "))
    frete = float(input("Digite o valor do frete: "))
    imposto1 = float(input("Digite o valor do Imposto 1 (%): "))
    imposto2 = float(input("Digite o valor do Imposto 2 (%): "))
    imposto3 = float(input("Digite o valor do Imposto 3 (%): "))
    
 
    valor_custo = calcular_valor_custo(valor, quantidade, frete, imposto1, imposto2, imposto3)
    
   
    valor_venda = valor_custo * 1.2  
    

    sheet.append([nome_produto, valor, quantidade, frete, imposto1, imposto2, imposto3, valor_custo, valor_venda])
    workbook.save("produtos.xlsx")
    
    print("Produto cadastrado com sucesso!")


def listar_produtos():
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("\nNome do Produto:", row[0])
        print("Valor:", row[1])
        print("Quantidade:", row[2])
        print("Frete:", row[3])
        print("Imposto 1 (%):", row[4])
        print("Imposto 2 (%):", row[5])
        print("Imposto 3 (%):", row[6])
        print("Valor de Custo:", row[7])
        print("Valor de Venda:", row[8])
        print("-------------------------")


def editar_produto():
    nome_produto = input("Digite o nome do produto que deseja editar: ")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == nome_produto:
            print("Dados do produto:")
            print(f"Nome do Produto: {row[0]}")
            print(f"Valor: {row[1]}")
            print(f"Quantidade: {row[2]}")
            print(f"Frete: {row[3]}")
            print(f"Imposto 1 (%): {row[4]}")
            print(f"Imposto 2 (%): {row[5]}")
            print(f"Imposto 3 (%): {row[6]}")
            print(f"Valor de Custo: {row[7]}")
            print(f"Valor de Venda: {row[8]}")
            
            opcao = input("Digite a opção que deseja editar (valor/quantidade/frete/imposto1/imposto2/imposto3): ").lower()
            novo_valor = input(f"Digite o novo valor para {opcao}: ")
            
            if opcao == "valor":
                row[1] = float(novo_valor)
            elif opcao == "quantidade":
                row[2] = int(novo_valor)
            elif opcao == "frete":
                row[3] = float(novo_valor)
            elif opcao == "imposto1":
                row[4] = float(novo_valor)
            elif opcao == "imposto2":
                row[5] = float(novo_valor)
            elif opcao == "imposto3":
                row[6] = float(novo_valor)
            
            # Recalcula o valor de custo e valor de venda após a edição
            row[7] = calcular_valor_custo(row[1], row[2], row[3], row[4], row[5], row[6])
            row[8] = row[7] * 1.2  # Pode ajustar a margem de lucro conforme necessário
            
            print("Produto editado com sucesso!")
            workbook.save("produtos.xlsx")
            return
    print("Produto não encontrado.")


def excluir_produto():
    nome_produto = input("Digite o nome do produto que deseja excluir: ")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == nome_produto:
            sheet.delete_rows(row[0].row)
            print("Produto excluído com sucesso!")
            workbook.save("produtos.xlsx")
            return
    print("Produto não encontrado.")

try:
    workbook = openpyxl.load_workbook("produtos.xlsx")
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nome do Produto", "Valor", "Quantidade", "Frete", "Imposto 1 (%)", "Imposto 2 (%)", "Imposto 3 (%)", "Valor de Custo", "Valor de Venda"])
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)
    sheet['D1'].font = Font(bold=True)
    sheet['E1'].font = Font(bold=True)
    sheet['F1'].font = Font(bold=True)
    sheet['G1'].font = Font(bold=True)
    sheet['H1'].font = Font(bold=True)
    sheet['I1'].font = Font(bold=True)
else:
    sheet = workbook.active

while True:
    print("\nMenu:")
    print("1. Cadastrar Produto")
    print("2. Editar Produto")
    print("3. Excluir Produto")
    print("4. Listar Produtos")
    print("5. Sair")

    opcao = input("Digite a opção desejada: ")

    if opcao == "1":
        cadastrar_produto()
    elif opcao == "2":
        editar_produto()
    elif opcao == "3":
        excluir_produto()
    elif opcao == "4":
        listar_produtos()
    elif opcao == "5":
        print("Saindo...")
        break
    else:
        print("Opção inválida. Tente novamente.")